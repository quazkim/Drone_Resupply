#!/usr/bin/env python3
import subprocess
import re
import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 15 instances from the table
instances = [
    ("Instances 2/U_10_0.5_Num_1.txt", 131),
    ("Instances 2/U_10_0.5_Num_2.txt", 160),
    ("Instances 2/U_10_0.5_Num_3.txt", 134),
    ("Instances 2/U_10_0.5_Num_4.txt", 150),
    ("Instances 2/U_10_0.5_Num_5.txt", 124),
    ("Instances 2/U_10_1.0_Num_1.txt", 173),
    ("Instances 2/U_10_1.0_Num_2.txt", 214),
    ("Instances 2/U_10_1.0_Num_3.txt", 193),
    ("Instances 2/U_10_1.0_Num_4.txt", 234),
    ("Instances 2/U_10_1.0_Num_5.txt", 175),
    ("Instances 2/U_10_1.5_Num_1.txt", 246),
    ("Instances 2/U_10_1.5_Num_2.txt", 199),
    ("Instances 2/U_10_1.5_Num_3.txt", 248),
    ("Instances 2/U_10_1.5_Num_4.txt", 350),
    ("Instances 2/U_10_1.5_Num_5.txt", 271),
]

print("="*80)
print("RUNNING 15 TEST INSTANCES FROM INSTANCES 2 (POPULATION = 200)")
print("="*80)
print()

results = []
better_count = 0
equal_count = 0
worse_count = 0
total_diff = 0.0
start_time = time.time()

for inst_path, paper_cost in instances:
    inst_name = os.path.basename(inst_path)
    print(f"Testing {inst_name:<35} (Paper: {paper_cost:4.0f} min)...", end=" ", flush=True)
    
    try:
        cmd = f'./main_ga_tabu "{inst_path}"'
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=300)
        
        cost_match = re.search(r'Cost before Local Search:\s*([\d.]+)', result.stdout)
        if cost_match:
            ga_cost = float(cost_match.group(1))
            diff = ga_cost - paper_cost
            diff_pct = (diff / paper_cost) * 100
            total_diff += diff_pct
            
            if diff < -0.1:
                better_count += 1
                status = "✓"
            elif diff <= 0.1:
                equal_count += 1
                status = "="
            else:
                worse_count += 1
                status = "✗"
            
            results.append((inst_name, paper_cost, ga_cost, diff_pct, status))
            print(f"GA+Tabu: {ga_cost:7.2f} | Diff: {diff_pct:+6.2f}% {status}")
        else:
            print("FAILED - Could not parse output")
    except subprocess.TimeoutExpired:
        print("TIMEOUT (>300s)")
    except Exception as e:
        print(f"ERROR: {e}")

elapsed = time.time() - start_time

print()
print("="*80)
print("SUMMARY TABLE (POPULATION = 200, Instances 2)")
print("="*80)
print(f"{'Instance':<35} {'Paper':<10} {'GA+Tabu':<10} {'Diff %':<10} {'Status':<10}")
print("-"*80)

for instance, paper, ga, diff_pct, status in results:
    status_name = "BETTER" if status == "✓" else ("EQUAL" if status == "=" else "WORSE")
    print(f"{instance:<35} {paper:<10.0f} {ga:<10.2f} {diff_pct:<+10.2f} {status_name:<10}")

print()
print(f"Better: {better_count}/15 ({100*better_count/15:.1f}%)")
print(f"Equal:  {equal_count}/15 ({100*equal_count/15:.1f}%)")
print(f"Worse:  {worse_count}/15 ({100*worse_count/15:.1f}%)")
print(f"Average difference: {total_diff/15:+.2f}% (GA+Tabu {'wins' if total_diff < 0 else 'loses'} on average)")
print(f"Total runtime: {elapsed/60:.1f} minutes")
print()

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Test Results"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
better_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
worse_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Add title
ws['A1'] = "GA+TABU TEST RESULTS - INSTANCES 2 (POPULATION = 200)"
ws['A1'].font = Font(bold=True, size=12)
ws.merge_cells('A1:E1')

# Add summary
ws['A3'] = "Summary"
ws['A3'].font = Font(bold=True)
ws['A4'] = f"Better: {better_count}/15 ({100*better_count/15:.1f}%)"
ws['A5'] = f"Worse: {worse_count}/15 ({100*worse_count/15:.1f}%)"
ws['A6'] = f"Equal: {equal_count}/15 ({100*equal_count/15:.1f}%)"
ws['A7'] = f"Average difference: {total_diff/15:+.2f}%"
ws['A8'] = f"Runtime: {elapsed/60:.1f} minutes"

# Add headers
headers = ["Instance", "Paper Cost (min)", "GA+Tabu Cost (min)", "Difference (%)", "Status"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=10, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Add data
for row_idx, (instance, paper, ga, diff_pct, status) in enumerate(results, 11):
    status_name = "BETTER" if status == "✓" else ("EQUAL" if status == "=" else "WORSE")
    
    ws.cell(row=row_idx, column=1).value = instance
    ws.cell(row=row_idx, column=2).value = paper
    ws.cell(row=row_idx, column=3).value = round(ga, 2)
    ws.cell(row=row_idx, column=4).value = round(diff_pct, 2)
    ws.cell(row=row_idx, column=5).value = status_name
    
    # Apply conditional formatting
    if status == "✓":
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).fill = better_fill
    elif status == "✗":
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).fill = worse_fill
    
    # Apply borders
    for col in range(1, 6):
        ws.cell(row=row_idx, column=col).border = border
    
    # Right align numbers
    for col in range(2, 6):
        ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="right")

# Adjust column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 15

# Save to Excel
excel_file = "test_results.xlsx"
wb.save(excel_file)
print(f"✓ Excel file saved: {excel_file}")
